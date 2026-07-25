"""Build Harry_Shenk_manual_payout_2026-06.xlsx from model.json.

Layout mirrors the user's `Harry_Shenk_manually_calculated.xlsx` sample: a 6-row
block per mentee (Amount / Month Start / Harry's Percentage / % of mo to be paid /
Harry's Pay / Assured Take-Home), with three traceability rows added above Amount.
Every computed cell is a live Excel formula.
"""
import json, datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

M = json.load(open('model.json'))
BLOCKS = [b for b in M['blocks'] if b['columns']]
EMPTY = [b for b in M['blocks'] if not b['columns']]
ADJ = M['adjustments']
DASH = M['dash_lines']
ALIAS = {'Isaiah Hursh': 'Ike Hursh'}
# Adjustments tab data rows — bounded ranges beat whole-column SUMIFs.
ADJ_FIRST, ADJ_LAST = 5, 4 + len(ADJ)

FONT = 'Arial'
BLUE = Font(name=FONT, size=10, color='0000FF')          # hardcoded input from CA
BLACK = Font(name=FONT, size=10)
GREEN = Font(name=FONT, size=10, color='008000')         # link to another sheet
BOLD = Font(name=FONT, size=10, bold=True)
NAME_F = Font(name=FONT, size=12, bold=True, color='1F3864')
H1 = Font(name=FONT, size=14, bold=True, color='1F3864')
H2 = Font(name=FONT, size=11, bold=True, color='1F3864')
WHITE_B = Font(name=FONT, size=10, bold=True, color='FFFFFF')

FILL_HDR = PatternFill('solid', fgColor='1F3864')
FILL_BAND = PatternFill('solid', fgColor='D9E2F3')
FILL_INPUT = PatternFill('solid', fgColor='FFFF00')       # user-editable
FILL_WARN = PatternFill('solid', fgColor='FFC7CE')        # variance / conflict
FILL_OK = PatternFill('solid', fgColor='C6EFCE')
FILL_NOTE = PatternFill('solid', fgColor='FFF2CC')
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = '$#,##0.00;($#,##0.00);-'
MONEY0 = '$#,##0;($#,##0);-'
PCT = '0.0%'
FRAC = '0.0000'
DATE = 'mmmm d, yyyy'

wb = Workbook()

# =====================================================================  README
ws = wb.active
ws.title = 'README'
ws.sheet_view.showGridLines = False
rows = [
    ('h1', "Harry Shenk — manual mentor-payout calculation"),
    ('sub', "Built 2026-07-25 from the Notion Mentees Database (roster), the HJG raw-data export "
            "(invoice dates + amounts), and reconciled against the dashboard's June-2026 payout build."),
    ('gap', ''),
    ('h2', "Headline"),
    # B6 = manual total, B10 = dashboard total on the June Reconciliation bridge.
    ('kv', ("Manual method, June 2026", "=='June Reconciliation'!B6")),
    ('kv', ("Dashboard paystub, June 2026 (reviewed)", "=='June Reconciliation'!B10")),
    ('kv', ("Variance", "==('June Reconciliation'!B6-'June Reconciliation'!B10)")),
    ('gap', ''),
    ('h2', "The method (your formulas, unchanged)"),
    ('t', "Amount               = tier price billed on the invoice's mentoring line, plus any adjustment you switch ON"),
    ('t', "Month Start          = the invoice date (CoachAccountable date_of)"),
    ('t', "Harry's Percentage   = the mentor revenue share for that calendar month"),
    ('t', "% of mo to be paid   = 1 - DAY(Month Start) / DAY(EOMONTH(Month Start, 0))"),
    ('t', "Harry's Pay          = Amount x Pct x %ofMo  +  (1 - %ofMo prior) x Amount prior x Pct prior"),
    ('t', "Assured Take-Home    = Amount x Pct"),
    ('gap', ''),
    ('h2', "Decisions you made 2026-07-25"),
    ('t', "1. Proration uses REAL days in the month (May = 31, June = 30), per your =1-DAY()/DAY(EOMONTH()) formula."),
    ('t', "2. Amount = tier price. Credits and discounts are listed on the Adjustments tab, defaulted OFF, and you"),
    ('t', "   can switch any of them ON with a Y/N toggle — the whole workbook recalculates."),
    ('t', "3. David Weaver's two separate May invoices get two columns; Josh Lehman's duplicated 4x line on one"),
    ('t', "   invoice is treated as a billing correction (one column, duplicate line OFF on the Adjustments tab)."),
    ('t', "4. Each mentee's timeline runs from their first mentoring invoice through one month past their last."),
    ('gap', ''),
    ('h2', "Assumptions I had to make (flag anything you disagree with)"),
    ('t', "A. Roster is the Notion export: 32 mentees list Harry Shenk as Mentor. See the Roster tab for the two"),
    ('t', "   mentees CoachAccountable assigns to Harry that Notion does not (Bryce Wenger, William Beachy)."),
    ('t', "B. Harry's revenue share ramps on HIS tenure, not each mentee's. Earliest CA engagement = 2024-05-20,"),
    ('t', "   so tenure month 2 (June 2024) = 50% and July 2024 onward = 60%. Only Allen Miller's first column is"),
    ('t', "   affected. Your sample sheet also used a flat 60% for established mentees."),
    ('t', "C. Pay-eligible revenue = 'MN Subscription | (Nx Month) Zoom Meetings', plus the legacy labels"),
    ('t', "   'One-on-One Mentoring', 'Monthly Mentoring Subscription' and 'JumpStart Your Freedom & One-on-One"),
    ('t', "   Mentoring'. JYF Supervised Progress, One-Time Setup Fee, standalone JumpStart, MT Engagement and the"),
    ('t', "   Marriage Encounter line are NOT mentor pay. Every excluded line is shown on Invoice Source Data."),
    ('t', "D. A month with no invoice gets a zero column so the prior month's remainder still lands there."),
    ('t', "E. Where an invoice's mentoring line was billed below tier (Joel Mast, $400 not $425) the billed figure"),
    ('t', "   is used as-is — there is no credit line to toggle."),
    ('gap', ''),
    ('h2', "Tabs"),
    ('t', "Mentee Blocks        — the calculation, one block per mentee, live formulas"),
    ('t', "Adjustments          — every credit / duplicate line, with the Y/N include toggle (THE editable tab)"),
    ('t', "Monthly Roll-up      — Harry's total pay per calendar month, per mentee"),
    ('t', "June Reconciliation  — manual vs the dashboard paystub, per mentee, with the causes priced out"),
    ('t', "Roster               — Notion vs CoachAccountable roster comparison"),
    ('t', "Invoice Source Data  — every invoice behind the numbers, including the lines that were excluded"),
    ('gap', ''),
    ('h2', "Colour key"),
    ('t', "Blue text = hardcoded figure taken from CoachAccountable.  Black = formula.  Green = link to another tab."),
    ('t', "Yellow fill = a cell you are meant to edit.  Red fill = a variance or conflict to look at."),
]
r = 1
for kind, val in rows:
    if kind == 'h1':
        ws.cell(r, 1, val).font = H1
    elif kind == 'h2':
        ws.cell(r, 1, val).font = H2
    elif kind == 'sub':
        c = ws.cell(r, 1, val); c.font = Font(name=FONT, size=10, italic=True, color='595959')
    elif kind == 'kv':
        ws.cell(r, 1, val[0]).font = BOLD
        c = ws.cell(r, 2, val[1][1:]); c.font = GREEN; c.number_format = MONEY; c.fill = FILL_BAND
    elif kind == 't':
        ws.cell(r, 1, val).font = BLACK
    r += 1
ws.column_dimensions['A'].width = 112
ws.column_dimensions['B'].width = 16

# =============================================================  MENTEE BLOCKS
mb = wb.create_sheet('Mentee Blocks')
mb.sheet_view.showGridLines = False
LBL = ['Month', 'Invoice #', 'Tier price billed', 'Adjustments applied', 'Amount',
       'Month Start', "Harry's Percentage", '% of mo to be paid', "Harry's Pay",
       'Assured Take-Home']
mb.cell(1, 1, "Harry Shenk — mentee-by-mentee payout calculation").font = H1
mb.cell(2, 1, "One block per mentee, oldest mentoring invoice first. Column widths hold up to 13 months; "
              "the Total column is at the far right.").font = Font(name=FONT, size=9, italic=True, color='595959')
MAXC = max(len(b['columns']) for b in BLOCKS)
TOTCOL = 2 + MAXC          # first month col = 2 (B)
RESCOL = TOTCOL + 1

mb.column_dimensions['A'].width = 22
for i in range(MAXC):
    mb.column_dimensions[CL(2 + i)].width = 15
mb.column_dimensions[CL(TOTCOL)].width = 14
mb.column_dimensions[CL(RESCOL)].width = 14

pay_ref = {}   # (mentee, ym) -> list of 'Mentee Blocks'!X99 refs
r = 4
for b in BLOCKS:
    cols = b['columns']
    n = len(cols)
    hdr = mb.cell(r, 1, b['notion_name'])
    hdr.font = NAME_F
    meta = (f"client {b['client_id']}  ·  Notion: {b['notion_status']}  ·  "
            f"{cols[-1]['tier'] if cols[-1]['tier'] != '—' else cols[-2]['tier']} current  ·  "
            f"{n} month column(s)")
    mb.cell(r, 2, meta).font = Font(name=FONT, size=9, italic=True, color='595959')
    for cc in range(1, RESCOL + 1):
        mb.cell(r, cc).fill = FILL_BAND
    r += 1
    r_month, r_inv, r_tier, r_adj, r_amt = r, r + 1, r + 2, r + 3, r + 4
    r_start, r_pct, r_frac, r_pay, r_ass = r + 5, r + 6, r + 7, r + 8, r + 9
    for i, lab in enumerate(LBL):
        c = mb.cell(r + i, 1, lab)
        c.font = BOLD if lab in ("Harry's Pay", 'Amount', 'Assured Take-Home') else BLACK
    for j, col in enumerate(cols):
        L = CL(2 + j)
        prev = CL(1 + j) if j else None
        mb.cell(r_month, 2 + j, col['label']).font = BOLD
        mb.cell(r_month, 2 + j).alignment = Alignment(horizontal='center')
        # invoice #
        c = mb.cell(r_inv, 2 + j, int(col['invoice_number']) if col['invoice_number'] else '— no invoice —')
        c.font = BLUE if col['invoice_number'] else Font(name=FONT, size=9, italic=True, color='808080')
        c.alignment = Alignment(horizontal='center')
        c.number_format = '0'
        # tier price (hardcoded from CA)
        c = mb.cell(r_tier, 2 + j, col['tier_price']); c.font = BLUE; c.number_format = MONEY0
        if col['label_coach'] and 'Harry' not in col['label_coach']:
            c.fill = FILL_NOTE
            c.comment = Comment(f"CoachAccountable line item reads \"{col['primary_item']}\" — the product is "
                                f"named for another coach, but Notion and CA both assign this mentee to Harry.",
                                'reconciliation')
        # adjustments pulled from the Adjustments tab
        if col['adj_keys']:
            mb.cell(r_adj, 2 + j,
                    f"=SUMIF(Adjustments!$C${ADJ_FIRST}:$C${ADJ_LAST},{L}{r_inv},"
                    f"Adjustments!$H${ADJ_FIRST}:$H${ADJ_LAST})").font = GREEN
        else:
            mb.cell(r_adj, 2 + j, 0).font = BLACK
        mb.cell(r_adj, 2 + j).number_format = MONEY
        # amount
        c = mb.cell(r_amt, 2 + j, f"={L}{r_tier}+{L}{r_adj}")
        c.font = BOLD; c.number_format = MONEY
        # month start
        c = mb.cell(r_start, 2 + j)
        if col['date']:
            c.value = dt.datetime.strptime(col['date'], '%Y-%m-%d')
            c.number_format = DATE; c.font = BLUE
        else:
            c.value = '—'; c.font = Font(name=FONT, size=9, italic=True, color='808080')
        c.alignment = Alignment(horizontal='center')
        # pct
        c = mb.cell(r_pct, 2 + j, col['pct']); c.number_format = PCT; c.font = BLUE
        c.alignment = Alignment(horizontal='center')
        # % of month to be paid — the user's formula
        c = mb.cell(r_frac, 2 + j)
        c.value = 0 if col['gap'] else f"=1-DAY({L}{r_start})/DAY(EOMONTH({L}{r_start},0))"
        c.number_format = FRAC; c.font = BLACK
        # Harry's Pay
        f = f"={L}{r_amt}*{L}{r_pct}*{L}{r_frac}"
        if prev:
            f += f"+(1-{prev}{r_frac})*{prev}{r_amt}*{prev}{r_pct}"
        c = mb.cell(r_pay, 2 + j, f); c.number_format = MONEY; c.font = BOLD
        # Assured
        c = mb.cell(r_ass, 2 + j, f"={L}{r_amt}*{L}{r_pct}"); c.number_format = MONEY
        pay_ref.setdefault((b['notion_name'], col['ym']), []).append(f"'Mentee Blocks'!{L}{r_pay}")
    last = CL(1 + n)
    tl, rl = CL(TOTCOL), CL(RESCOL)
    mb.cell(r_month, TOTCOL, 'TOTAL').font = BOLD
    mb.cell(r_month, RESCOL, 'Residual').font = BOLD
    for rr in (r_pay, r_ass):
        c = mb.cell(rr, TOTCOL, f"=SUM(B{rr}:{last}{rr})"); c.number_format = MONEY; c.font = BOLD
    c = mb.cell(r_ass, RESCOL, f"={tl}{r_ass}-{tl}{r_pay}")
    c.number_format = MONEY; c.font = BOLD
    c.comment = Comment("Assured Take-Home minus Harry's Pay: the roll-forward lag still outstanding "
                        "at the end of this timeline.", 'reconciliation')
    for cc in range(1, RESCOL + 1):
        mb.cell(r_ass, cc).border = Border(bottom=THIN)
    r = r_ass + 3

# mentees with no mentoring revenue
mb.cell(r, 1, "Notion mentees of Harry with no mentoring invoice in the raw data").font = H2
r += 1
mb.cell(r, 1, "These are JYF / waiting-list or not-yet-started mentees. No mentoring revenue means no payout line.")\
  .font = Font(name=FONT, size=9, italic=True, color='595959')
r += 1
for b in EMPTY:
    mb.cell(r, 1, b['notion_name']).font = BLACK
    mb.cell(r, 2, f"client {b['client_id']}  ·  Notion: {b['notion_status']}").font = \
        Font(name=FONT, size=9, italic=True, color='595959')
    r += 1

# ===============================================================  ADJUSTMENTS
aw = wb.create_sheet('Adjustments')
aw.sheet_view.showGridLines = False
aw.cell(1, 1, "Adjustments — credits, discounts and duplicate lines").font = H1
aw.cell(2, 1, "Switch Include? to Y to let a line reduce (or raise) the pay basis. Everything defaults to N, "
              "which is 'pay Harry on the tier price'. The Mentee Blocks tab recalculates automatically.")\
  .font = Font(name=FONT, size=10, italic=True, color='595959')
hdrs = ['Key', 'Mentee', 'Invoice #', 'Invoice date', 'Line item', 'Amount', 'Include?', 'Effective']
for i, h in enumerate(hdrs):
    c = aw.cell(4, 1 + i, h); c.font = WHITE_B; c.fill = FILL_HDR; c.border = BOX
    c.alignment = Alignment(horizontal='center')
name_by_id = {b['client_id']: b['notion_name'] for b in M['blocks']}
for i, a in enumerate(sorted(ADJ, key=lambda z: (name_by_id.get(z['client_id'], ''), z['date']))):
    rr = 5 + i
    aw.cell(rr, 1, a['key']).font = BLACK
    aw.cell(rr, 2, name_by_id.get(a['client_id'], str(a['client_id']))).font = BLACK
    c = aw.cell(rr, 3, int(a['invoice_number'])); c.font = BLUE; c.number_format = '0'
    c = aw.cell(rr, 4, dt.datetime.strptime(a['date'], '%Y-%m-%d')); c.number_format = DATE; c.font = BLUE
    aw.cell(rr, 5, f"{a['item']}   [{a['kind']}]").font = BLACK
    c = aw.cell(rr, 6, a['amount']); c.number_format = MONEY; c.font = BLUE
    c = aw.cell(rr, 7, a['default_include'])
    c.font = Font(name=FONT, size=10, bold=True, color='0000FF'); c.fill = FILL_INPUT
    c.alignment = Alignment(horizontal='center'); c.border = BOX
    c = aw.cell(rr, 8, f"=IF(G{rr}=\"Y\",F{rr},0)"); c.number_format = MONEY; c.font = BLACK
lastadj = 4 + len(ADJ)
dv = DataValidation(type='list', formula1='"Y,N"', allow_blank=False, showDropDown=False)
aw.add_data_validation(dv)
dv.add(f'G5:G{lastadj}')
aw.cell(lastadj + 2, 5, 'Total adjustment currently applied').font = BOLD
c = aw.cell(lastadj + 2, 8, f"=SUM(H5:H{lastadj})"); c.number_format = MONEY; c.font = BOLD
aw.cell(lastadj + 4, 1, "Note: the dashboard's June build had a reviewer remove exactly these credits, which is why "
                        "its 'Engine payout' ($3,017.70) and 'Effective payout' ($3,273.50) differ.")\
  .font = Font(name=FONT, size=9, italic=True, color='595959')
for col, w in zip('ABCDEFGH', (12, 20, 11, 18, 68, 12, 10, 12)):
    aw.column_dimensions[col].width = w

# ============================================================  MONTHLY ROLLUP
mr = wb.create_sheet('Monthly Roll-up')
mr.sheet_view.showGridLines = False
mr.cell(1, 1, "Harry Shenk — total pay by calendar month (manual method)").font = H1
mr.cell(2, 1, "Each cell links straight to that mentee's Harry's Pay cell on the Mentee Blocks tab.")\
  .font = Font(name=FONT, size=10, italic=True, color='595959')
all_ym = sorted({c['ym'] for b in BLOCKS for c in b['columns']})
c = mr.cell(4, 1, 'Mentee'); c.font = WHITE_B; c.fill = FILL_HDR; c.border = BOX
for j, ym in enumerate(all_ym):
    c = mr.cell(4, 2 + j, ym); c.font = WHITE_B; c.fill = FILL_HDR; c.border = BOX
    c.alignment = Alignment(horizontal='center')
c = mr.cell(4, 2 + len(all_ym), 'TOTAL'); c.font = WHITE_B; c.fill = FILL_HDR; c.border = BOX
JUNE_ROW = {}
for i, b in enumerate(BLOCKS):
    rr = 5 + i
    mr.cell(rr, 1, b['notion_name']).font = BLACK
    for j, ym in enumerate(all_ym):
        refs = pay_ref.get((b['notion_name'], ym))
        cc = mr.cell(rr, 2 + j)
        if refs:
            cc.value = '=' + '+'.join(refs); cc.font = GREEN
        else:
            cc.value = 0; cc.font = Font(name=FONT, size=10, color='BFBFBF')
        cc.number_format = MONEY
        if ym == '2026-06':
            JUNE_ROW[b['notion_name']] = f"'Monthly Roll-up'!{CL(2 + j)}{rr}"
    tc = mr.cell(rr, 2 + len(all_ym), f"=SUM(B{rr}:{CL(1 + len(all_ym))}{rr})")
    tc.number_format = MONEY; tc.font = BOLD
tr = 5 + len(BLOCKS)
mr.cell(tr, 1, 'TOTAL').font = BOLD
for j in range(len(all_ym) + 1):
    c = mr.cell(tr, 2 + j, f"={CL(2 + j)}5:{CL(2 + j)}{tr - 1}")
    c.value = f"=SUM({CL(2 + j)}5:{CL(2 + j)}{tr - 1})"
    c.number_format = MONEY; c.font = BOLD; c.border = Border(top=THIN)
mr.column_dimensions['A'].width = 22
for j in range(len(all_ym) + 1):
    mr.column_dimensions[CL(2 + j)].width = 12
mr.freeze_panes = 'B5'

# ========================================================  JUNE RECONCILIATION
jr = wb.create_sheet('June Reconciliation')
jr.sheet_view.showGridLines = False
jr.cell(1, 1, "June 2026 — manual calculation vs the dashboard paystub").font = H1
jr.cell(2, 1, "Manual figures link live to the Monthly Roll-up. Dashboard figures are typed from "
              "payoutbuildharryshenk20260620260725.csv.").font = Font(name=FONT, size=10, italic=True, color='595959')

jr.cell(4, 1, 'Bridge — what explains the gap').font = H2
# Labels must not begin with "=" or "+" — Excel would read them as formulas.
bridge = [
    ("Manual method as built (this workbook)", None, ''),
    ("Step 1 — proration on a fixed 30-day month instead of real days", 51.9064516129,
     "The dashboard divides by 30 always; your formula divides by 31 in May."),
    ("Step 2 — David Weaver's two May invoices grouped into one calendar month", 144.50,
     "Chained columns send the 17 May remainder into the 21 May column; the dashboard sends both into June."),
    ("Step 3 — Josh Lehman's duplicate 4x line counted as revenue", 93.50,
     "You chose to treat it as a billing correction; the dashboard pays on it."),
    ("Equals — dashboard paystub, June 2026 (Effective payout)", None, ''),
]
for i, h in enumerate(['Step', 'Amount', 'Why']):
    c = jr.cell(5, 1 + i, h); c.font = WHITE_B; c.fill = FILL_HDR; c.border = BOX
for i, (lab, amt, why) in enumerate(bridge):
    rr = 6 + i
    jr.cell(rr, 1, lab).font = BOLD if amt is None else BLACK
    c = jr.cell(rr, 2)
    if amt is None:
        c.value = None
    else:
        c.value = amt; c.font = BLUE
    c.number_format = MONEY
    jr.cell(rr, 3, why).font = Font(name=FONT, size=9, color='595959')
DETAIL_TOP = 14
jr['B6'] = f"=SUM(B{DETAIL_TOP}:B{DETAIL_TOP + len(DASH) - 1})"; jr['B6'].font = BOLD
jr['B10'] = "=B6+B7+B8+B9"; jr['B10'].font = BOLD
jr['B10'].fill = FILL_BAND
jr['B6'].fill = FILL_BAND
jr['B10'].number_format = MONEY
# Proof the bridge lands exactly on the paystub total from the detail table below.
c = jr.cell(10, 3, f"=B10-C{DETAIL_TOP + len(DASH)}")
c.number_format = MONEY; c.font = BOLD
jr.cell(10, 4, "<- bridge minus the paystub total below; must be zero")\
  .font = Font(name=FONT, size=9, italic=True, color='595959')
jr.cell(12, 1, 'Line by line').font = H2
for i, h in enumerate(['Mentee', 'Manual (this workbook)', 'Dashboard — Effective',
                       'Variance vs Effective', 'Dashboard — Engine (pre-review)',
                       'Variance vs Engine', 'What differs']):
    c = jr.cell(13, 1 + i, h); c.font = WHITE_B; c.fill = FILL_HDR; c.border = BOX
    c.alignment = Alignment(horizontal='center', wrap_text=True)

CAUSE = {
    'David Weaver': "Two May invoices (4108 on the 17th, 4113 on the 21st). Your two-column layout keeps the "
                    "17 May remainder inside May; the dashboard rolls both May invoices into June. Plus 31-vs-30 days.",
    'Josh Lehman': "Invoice 4109 carries two $425 4x lines. You treat the second as a correction (toggle OFF on "
                   "Adjustments); the dashboard pays on both. Plus 31-vs-30 days.",
    'Gordon Boone': "Matches — first mentoring invoice is 17 June, and June has 30 days so both methods agree.",
    'Nelson Miller': "Matches — 11 June invoice, no May mentoring invoice to roll forward, 30-day month.",
    'Russ Zehr': "Matches — 29 June invoice, no May mentoring invoice, 30-day month.",
}
DEFAULT_CAUSE = ("Proration denominator only: your formula divides the May invoice by 31 real days, "
                 "the dashboard divides by a fixed 30.")
june_names = sorted(DASH, key=lambda k: -DASH[k]['effective'])
inv_alias = {v: k for k, v in ALIAS.items()}
for i, dn in enumerate(june_names):
    rr = DETAIL_TOP + i
    mine = inv_alias.get(dn, dn)
    jr.cell(rr, 1, dn).font = BLACK
    ref = JUNE_ROW.get(mine)
    c = jr.cell(rr, 2, f"={ref}" if ref else 0); c.number_format = MONEY; c.font = GREEN
    c = jr.cell(rr, 3, DASH[dn]['effective']); c.number_format = MONEY; c.font = BLUE
    c = jr.cell(rr, 4, f"=B{rr}-C{rr}"); c.number_format = MONEY; c.font = BOLD
    c = jr.cell(rr, 5, DASH[dn]['engine']); c.number_format = MONEY; c.font = BLUE
    c = jr.cell(rr, 6, f"=B{rr}-E{rr}"); c.number_format = MONEY
    cause = CAUSE.get(mine, CAUSE.get(dn, DEFAULT_CAUSE))
    jr.cell(rr, 7, cause).font = Font(name=FONT, size=9, color='595959')
    jr.cell(rr, 7).alignment = Alignment(wrap_text=True, vertical='top')
    fill = FILL_OK if abs(DASH[dn]['effective'] - DASH[dn]['engine']) < 0.005 else FILL_NOTE
    jr.cell(rr, 5).fill = fill
TR = DETAIL_TOP + len(june_names)
jr.cell(TR, 1, 'TOTAL').font = BOLD
for col in 'BCDEF':
    c = jr.cell(TR, 'ABCDEFG'.index(col) + 1, f"=SUM({col}{DETAIL_TOP}:{col}{TR - 1})")
    c.number_format = MONEY; c.font = BOLD; c.border = Border(top=THIN)
jr.cell(TR + 2, 1, "Engine vs Effective: the dashboard's raw engine number is what it computes from the net invoice "
                   "(credits applied). A reviewer removed six credit lines, which is the $255.80 difference. "
                   "This workbook defaults to the same position — tier price, credits OFF.")\
  .font = Font(name=FONT, size=9, italic=True, color='595959')
jr.cell(TR + 3, 1, "Amber fill in the Engine column = a line the reviewer overrode.")\
  .font = Font(name=FONT, size=9, italic=True, color='595959')
for col, w in zip('ABCDEFG', (22, 20, 20, 20, 22, 20, 78)):
    jr.column_dimensions[col].width = w
jr.column_dimensions['C'].width = 20

# ====================================================================  ROSTER
rs = wb.create_sheet('Roster')
rs.sheet_view.showGridLines = False
rs.cell(1, 1, "Roster — Notion (source of truth) vs CoachAccountable").font = H1
rs.cell(2, 1, "The Notion Mentees Database is authoritative for who mentors whom. This tab shows where "
              "CoachAccountable disagrees.").font = Font(name=FONT, size=10, italic=True, color='595959')
mdf = pd.read_csv('raw/mentees.csv')
mrow = {int(r['client_id']): r for _, r in mdf.iterrows() if pd.notna(r['client_id'])}
for i, h in enumerate(['Notion mentee', 'Notion status', 'Notion mentor', 'CA client ID', 'CA name',
                       'CA owner coach', 'CA tier', 'Mentoring months in data', 'In June 2026 payout?', 'Flag']):
    c = rs.cell(4, 1 + i, h); c.font = WHITE_B; c.fill = FILL_HDR; c.border = BOX
    c.alignment = Alignment(horizontal='center', wrap_text=True)
allb = sorted(M['blocks'], key=lambda b: b['notion_name'])
rr = 5
for b in allb:
    cid = b['client_id']
    ca = mrow.get(cid) if cid else None
    owner = ca['ca_owner_coach_name'] if ca is not None and pd.notna(ca['ca_owner_coach_name']) else '—'
    nmonths = len([c for c in b['columns'] if not c['gap']])
    injune = 'yes' if any(c['ym'] == '2026-06' for c in b['columns']) else 'no'
    flag = ''
    if owner != 'Harry Shenk':
        flag = f'CA owner is {owner}'
    if ca is not None and ca.get('notion_coach_conflict') == True:
        flag = (flag + '; ' if flag else '') + \
               f"dashboard has a stored Notion coach conflict (it holds '{ca['notion_coach']}')"
    vals = [b['notion_name'], b['notion_status'], 'Harry Shenk', cid,
            ca['ca_name'] if ca is not None else '—', owner,
            ca['ca_current_tier'] if ca is not None else '—', nmonths, injune, flag]
    for j, v in enumerate(vals):
        c = rs.cell(rr, 1 + j, v); c.font = BLACK
        if j == 9 and v:
            c.fill = FILL_WARN
        if j == 8 and v == 'yes':
            c.fill = FILL_OK
    rr += 1
rr += 2
rs.cell(rr, 1, "Assigned to Harry in CoachAccountable but NOT in the Notion export").font = H2
rr += 1
for i, h in enumerate(['CA name', 'CA client ID', 'CA tier', 'CA status', 'Notion mentor', 'Notion status',
                       'June 2026 impact']):
    c = rs.cell(rr, 1 + i, h); c.font = WHITE_B; c.fill = FILL_HDR; c.border = BOX
rr += 1
extras = [
    ('Bryce Wenger', 301320, '4x', 'active', '~None Assigned', 'Waiting List (JYF)',
     'none — his 24 June invoice is JYF only. But his 23 July invoice is a $425 4x line, so he WILL '
     'matter for the July payout. Decide who owns him before then.'),
    ('William Beachy', 252530, '2x', 'inactive', '(blank)', 'Done (Quit OR No Mentoring)',
     'none — no 2026 invoices at all.'),
]
for e in extras:
    for j, v in enumerate(e):
        c = rs.cell(rr, 1 + j, v); c.font = BLACK
        if j == 6:
            c.fill = FILL_WARN; c.alignment = Alignment(wrap_text=True, vertical='top')
    rr += 1
for col, w in zip('ABCDEFGHIJ', (24, 28, 16, 13, 22, 20, 11, 15, 16, 62)):
    rs.column_dimensions[col].width = w

# ======================================================  INVOICE SOURCE DATA
sd = wb.create_sheet('Invoice Source Data')
sd.sheet_view.showGridLines = False
sd.cell(1, 1, "Invoice source data — every mentoring invoice behind this workbook").font = H1
sd.cell(2, 1, "From the hjgrawdata 2026-07-25 export, table ca_invoices. 'Lines excluded' shows what was on the "
              "invoice but is not mentor pay.").font = Font(name=FONT, size=10, italic=True, color='595959')
for i, h in enumerate(['Mentee', 'Client ID', 'Invoice #', 'Invoice date', 'Mentoring line item', 'Tier',
                       'Tier price', 'Invoice total billed', 'Collected', 'Lines excluded from pay',
                       'Adjustment lines (see Adjustments tab)']):
    c = sd.cell(4, 1 + i, h); c.font = WHITE_B; c.fill = FILL_HDR; c.border = BOX
    c.alignment = Alignment(horizontal='center', wrap_text=True)
adj_by_key = {a['key']: a for a in ADJ}
rr = 5
for b in sorted(BLOCKS, key=lambda z: z['notion_name']):
    for col in b['columns']:
        if col['gap']:
            continue
        exc = '; '.join(f"{it} (${am:,.2f})" for it, am in col['excluded_lines']) or '—'
        adjs = '; '.join(f"{adj_by_key[k]['item']} (${adj_by_key[k]['amount']:,.2f})" for k in col['adj_keys']) or '—'
        vals = [b['notion_name'], b['client_id'], int(col['invoice_number']),
                dt.datetime.strptime(col['date'], '%Y-%m-%d'), col['primary_item'], col['tier'],
                col['tier_price'], col['invoice_amount'], col['collected'], exc, adjs]
        for j, v in enumerate(vals):
            c = sd.cell(rr, 1 + j, v); c.font = BLACK
            if j == 3:
                c.number_format = DATE
            if j in (6, 7, 8):
                c.number_format = MONEY
        if col['collected'] < col['invoice_amount'] - 0.005:
            sd.cell(rr, 9).fill = FILL_WARN
        if adjs != '—':
            sd.cell(rr, 11).fill = FILL_NOTE
        rr += 1
sd.cell(rr + 1, 1, "Pink fill in Collected = billed but not fully paid. Mentor pay is calculated on BILLED "
                   "revenue, matching both your sample sheet and the dashboard.")\
  .font = Font(name=FONT, size=9, italic=True, color='595959')
for col, w in zip('ABCDEFGHIJK', (22, 11, 11, 18, 56, 8, 12, 16, 12, 52, 52)):
    sd.column_dimensions[col].width = w
sd.freeze_panes = 'A5'

# global font pass for anything left at default
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for c in row:
            if c.value is not None and (c.font is None or c.font.name != FONT):
                f = c.font
                c.font = Font(name=FONT, size=f.size or 10, bold=f.bold, italic=f.italic,
                              color=f.color)

OUT = 'Harry_Shenk_manual_payout_2026-06.xlsx'
wb.save(OUT)
print('wrote', OUT)
